#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
xlsx_to_pubtables_json.py

Конвертация XLSX в JSON в формате, близком к PubTables-1M.

Поддержка:
- обход всех .xlsx в папке рекурсивно;
- создание отдельной папки для JSON;
- обработка объединённых ячеек;
- распознавание цветных ячеек:
    * жёлтый  -> is_column_header = True
    * зелёный  -> is_projected_row_header = True
    * синий   -> is_metadata = True
- логирование по каждому файлу;
- сохранение относительной структуры подпапок.

Зависимости:
    pip install openpyxl
"""

from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional, Dict, List, Tuple

from openpyxl import load_workbook


# =========================
# НАСТРОЙКИ ПУТЕЙ
# =========================

# Папка с исходными XLSX
INPUT_DIR = Path(r"c:\Users\Юзя\Desktop\BetterExperimentAmpere\Convert_from_xlsx_to_Json\table_normalization")

# Папка для выходных JSON
OUTPUT_DIR = INPUT_DIR.parent / f"{INPUT_DIR.name}_converted_json"

# Папка для логов
LOG_DIR = OUTPUT_DIR / "_logs"

# Обрабатывать только активный лист
USE_ACTIVE_SHEET_ONLY = True


# =========================
# ЦВЕТА
# =========================

YELLOW = {"FFFF00", "FFFFFF00"}
GREEN = {"00B050", "FF00B050"}
BLUE = {"00B0F0", "FF00B0F0"}


# =========================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# =========================

def normalize_rgb(value: Any) -> Optional[str]:
    """
    Приводит цвет к верхнему регистру.
    Поддерживает 6- и 8-значный RGB/ARGB.
    """
    if value is None:
        return None

    s = str(value).strip().upper()
    if not s:
        return None

    return s


def get_cell_fill_rgb(cell) -> Optional[str]:
    """
    Возвращает RGB/ARGB цвет заливки ячейки, если он задан.
    Иначе None.
    """
    try:
        fg = cell.fill.fgColor
        if fg is None:
            return None

        # Самый частый случай: fgColor.type == "rgb"
        if getattr(fg, "type", None) == "rgb" and getattr(fg, "rgb", None):
            return normalize_rgb(fg.rgb)

        # Иногда значение лежит в fgColor.value
        if getattr(fg, "value", None):
            return normalize_rgb(fg.value)

    except Exception:
        pass

    return None


def safe_text(value: Any) -> str:
    """
    Приводит значение ячейки к строке для JSON.
    """
    if value is None:
        return ""
    return str(value).strip()


@dataclass
class MergedRange:
    min_row: int
    min_col: int
    max_row: int
    max_col: int


def build_merged_map(sheet) -> Dict[Tuple[int, int], MergedRange]:
    """
    Строит карту всех ячеек, входящих в merged range.
    """
    merged_map: Dict[Tuple[int, int], MergedRange] = {}

    for rng in sheet.merged_cells.ranges:
        mr = MergedRange(
            min_row=rng.min_row,
            min_col=rng.min_col,
            max_row=rng.max_row,
            max_col=rng.max_col,
        )

        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merged_map[(r, c)] = mr

    return merged_map


def classify_by_color(fill_rgb: Optional[str]) -> Tuple[bool, bool, bool]:
    """
    Возвращает:
        (is_column_header, is_projected_row_header, is_metadata)
    """
    is_col_header = False
    is_row_header = False
    is_metadata = False

    if not fill_rgb:
        return is_col_header, is_row_header, is_metadata

    if fill_rgb in YELLOW:
        is_col_header = True
    elif fill_rgb in GREEN:
        is_row_header = True
    elif fill_rgb in BLUE:
        is_metadata = True

    return is_col_header, is_row_header, is_metadata


# =========================
# ПАРСИНГ ОДНОГО ФАЙЛА
# =========================

def parse_xlsx_to_pubtables(
    xlsx_path: Path,
    log_dir: Optional[Path] = None,
) -> Optional[dict]:
    """
    Конвертирует один XLSX в словарь формата PubTables-1M.
    Возвращает None, если файл не удалось обработать.
    """
    try:
        wb = load_workbook(xlsx_path, data_only=True)
        sheet = wb.active if USE_ACTIVE_SHEET_ONLY else wb[wb.sheetnames[0]]
    except Exception as e:
        if log_dir:
            log_dir.mkdir(parents=True, exist_ok=True)
            log_file = log_dir / f"{xlsx_path.stem}.log"
            with open(log_file, "a", encoding="utf-8") as lf:
                lf.write(
                    f"=== ERROR loading {xlsx_path.name} ===\n"
                    f"{type(e).__name__}: {e}\n\n"
                )
        print(f"[SKIP] Не удалось открыть файл: {xlsx_path}")
        return None

    merged_map = build_merged_map(sheet)
    seen = set()
    cells_out: List[dict] = []
    log_lines: List[str] = []

    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            if (r, c) in seen:
                continue

            cell = sheet.cell(row=r, column=c)
            text = safe_text(cell.value)

            # Пропускаем полностью пустые ячейки
            if text == "":
                continue

            fill_rgb = get_cell_fill_rgb(cell)
            is_col_header, is_row_header, is_metadata = classify_by_color(fill_rgb)

            # A1 — как правило, и заголовок строки, и заголовок столбца
            if r == 1 and c == 1:
                is_col_header = True
                is_row_header = True

            # Объединённые ячейки
            if (r, c) in merged_map:
                mr = merged_map[(r, c)]
                row_nums = list(range(mr.min_row - 1, mr.max_row))
                col_nums = list(range(mr.min_col - 1, mr.max_col))

                for rr in range(mr.min_row, mr.max_row + 1):
                    for cc in range(mr.min_col, mr.max_col + 1):
                        seen.add((rr, cc))
            else:
                row_nums = [r - 1]
                col_nums = [c - 1]

            log_lines.append(
                f"Cell {cell.coordinate} | "
                f"value='{text}' | "
                f"fill={fill_rgb} | "
                f"rows={row_nums} | cols={col_nums} | "
                f"col_header={is_col_header} | "
                f"row_header={is_row_header} | "
                f"metadata={is_metadata}"
            )

            cells_out.append(
                {
                    "row_nums": row_nums,
                    "column_nums": col_nums,
                    "xml_text_content": text,
                    "pdf_text_content": text,
                    "is_column_header": is_col_header,
                    "is_projected_row_header": is_row_header,
                    "is_metadata": is_metadata,
                }
            )

    if log_dir:
        log_dir.mkdir(parents=True, exist_ok=True)
        log_file = log_dir / f"{xlsx_path.stem}.log"
        with open(log_file, "a", encoding="utf-8") as lf:
            lf.write(f"=== {xlsx_path.name} ===\n")
            lf.write("\n".join(log_lines))
            lf.write("\n\n")

    return {
        "cells": cells_out,
        "structure_id": xlsx_path.stem,
        "pdf_file_name": f"{xlsx_path.stem}.png",
        "split": "manual_colored",
    }


# =========================
# ОБРАБОТКА ВСЕЙ ПАПКИ
# =========================

def process_all_files() -> None:
    """
    Рекурсивно проходит по всем .xlsx в INPUT_DIR и сохраняет JSON в OUTPUT_DIR,
    сохраняя относительную структуру подпапок.
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    LOG_DIR.mkdir(parents=True, exist_ok=True)

    xlsx_files = sorted(
        p for p in INPUT_DIR.rglob("*.xlsx")
        if not p.name.startswith("~$")
    )

    summary_log = OUTPUT_DIR / "summary_log.txt"

    total = 0
    ok = 0
    failed = 0

    with open(summary_log, "w", encoding="utf-8") as sf:
        sf.write(f"INPUT_DIR: {INPUT_DIR}\n")
        sf.write(f"OUTPUT_DIR: {OUTPUT_DIR}\n")
        sf.write(f"FILES_FOUND: {len(xlsx_files)}\n\n")

    for xlsx_path in xlsx_files:
        total += 1
        rel_path = xlsx_path.relative_to(INPUT_DIR)
        out_json = OUTPUT_DIR / rel_path.with_suffix(".json")
        out_json.parent.mkdir(parents=True, exist_ok=True)

        table = parse_xlsx_to_pubtables(xlsx_path, log_dir=LOG_DIR)

        if table is None:
            failed += 1
            with open(summary_log, "a", encoding="utf-8") as sf:
                sf.write(f"[FAIL] {rel_path}\n")
            continue

        try:
            with open(out_json, "w", encoding="utf-8") as jf:
                json.dump([table], jf, ensure_ascii=False, indent=2)

            ok += 1
            print(f"[OK] {rel_path} -> {out_json.relative_to(OUTPUT_DIR)}")

            with open(summary_log, "a", encoding="utf-8") as sf:
                sf.write(f"[OK] {rel_path} -> {out_json.relative_to(OUTPUT_DIR)}\n")

        except Exception as e:
            failed += 1
            print(f"[ERROR] Не удалось сохранить JSON для {xlsx_path}: {e}")
            with open(summary_log, "a", encoding="utf-8") as sf:
                sf.write(f"[ERROR] {rel_path} | {type(e).__name__}: {e}\n")

    with open(summary_log, "a", encoding="utf-8") as sf:
        sf.write("\n=== DONE ===\n")
        sf.write(f"TOTAL: {total}\n")
        sf.write(f"OK: {ok}\n")
        sf.write(f"FAILED: {failed}\n")

    print("\nГотово.")
    print(f"Всего файлов: {total}")
    print(f"Успешно: {ok}")
    print(f"Ошибок: {failed}")
    print(f"JSON: {OUTPUT_DIR}")
    print(f"Логи: {LOG_DIR}")


# =========================
# MAIN
# =========================

if __name__ == "__main__":
    process_all_files()